import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
from openpyxl import Workbook
from tempfile import NamedTemporaryFile

st.set_page_config(layout="wide")

#logo
st.image("https://uploads-ssl.webflow.com/643e8b8f7e656b61bd29c098/644240807d96bd6b2322d3a0_OmeatFooterLogo.png",
         width=500)

st.markdown("<h1 style='text-align:center;'> Omeat YSI Analyzer </h1>", unsafe_allow_html=True)

num_plates = st.slider('How many plates are you going to upload?', 0, 100)

all_active_cells = []
all_sample_names = []
bio_files = []
ise_files = []

columns_ui = st.columns(3)
columns = ['A', 'B', 'C']
plate_order_dict = [f'R24_{ch}{str(num).zfill(2)}' for ch in columns for num in range(1, 9)]

# Functions to clean and summarize data
def clean_df(df, well_id):
    return df[['Chemistry', 'Concentration', 'Well Id']][df['Well Id'] == well_id]

def ysi_summary(df):
    df['Concentration'] = df['Concentration'].to_numeric()
    mean_cleaned = df.groupby('Chemistry').mean().round(3).reset_index().rename(columns={'Concentration': 'Mean Concentration (g/L)'})
    std_cleaned = df.groupby('Chemistry').std().round(3).reset_index().rename(columns={'Concentration': 'std (g/L)'}).drop(columns={'Chemistry'})
    return pd.concat([mean_cleaned, std_cleaned], axis=1)

for plate in range(num_plates):
    st.markdown(f"### Input for Plate {plate + 1}")
    YSI_cells = []

    # Create the 3 columns for the plate inputs
    with st.container():
        columns_ui = st.columns(3)
        for idx, col_name in enumerate(columns):
            with columns_ui[idx]:
                st.markdown(f"<h2 style='text-align:center;'> {col_name} </h2>", unsafe_allow_html=True)
                for i in range(1, 9):
                    cell_name = st.text_input(f'Plate {plate + 1} - {col_name}{i}', placeholder=f'{col_name}{i}', label_visibility='collapsed')
                    YSI_cells.append(cell_name)

    # Create a dictionary of active cells and store it
    active_cells = {plate_order_dict[idx]: sample for idx, sample in enumerate(YSI_cells) if sample}
    all_active_cells.append(active_cells)

    # Get the sample name
    sample_name = st.text_input(f"Sample name for Plate {plate + 1}")
    all_sample_names.append(sample_name)

    # Get the Bioanalysis file
    bio_file = st.file_uploader(f"**Upload your *Bioanalysis* file for Plate {plate + 1}!**")
    if bio_file:
        bio_files.append(pd.read_csv(bio_file, encoding='ISO-8859-1'))
    else:
        bio_files.append(None)

    # Get the ISEAnalysis file
    ise_file = st.file_uploader(f"**Upload your *ISEAnalysis* file for Plate {plate + 1} if you have one!**")
    if ise_file:
        ise_files.append(pd.read_csv(ise_file, encoding='ISO-8859-1'))
    else:
        ise_files.append(None)


# Create a dataframe to aggregate results
all_data = []

for plate in range(num_plates):
    pandas_bio_df = bio_files[plate]
    pandas_ise = ise_files[plate]
    active_cells = all_active_cells[plate]
    sample_name = all_sample_names[plate]

    if pandas_bio_df is not None:
        for i in active_cells:
            bio_df = ysi_summary(clean_df(pandas_bio_df, i))

            if pandas_ise is not None:
                ise_df = ysi_summary(clean_df(pandas_ise, i))
                combined_df = pd.concat([bio_df, ise_df])
            else:
                combined_df = bio_df

            combined_df['Experiment'] = sample_name
            combined_df['Well'] = active_cells[i]
            all_data.append(combined_df)

# Convert list of dataframes to a single dataframe
if all_data:
    df_all = pd.concat(all_data)

# Create faceted bar chart using plotly.express
    fig = px.bar(df_all,x='Well', y='Mean Concentration (g/L)', color = 'Chemistry', template= 'ggplot2',
                error_y='std (g/L)', facet_col='Chemistry', facet_col_wrap = 3, height = 800,
                title='YSI Analyzer Results', facet_row_spacing = .15, facet_col_spacing = .06,
                labels={'Mean Concentration (g/L)': 'Concentration'},
                hover_name='Well')

    fig.for_each_annotation(lambda a: a.update(text=a.text.split("=")[-1]))
    fig.update_xaxes(matches='x', title = '')
    fig.for_each_yaxis(lambda yaxis: yaxis.update(showticklabels=True))
    fig.for_each_xaxis(lambda xaxis: xaxis.update(showticklabels=True))
    fig.update_yaxes(matches=None, title = '' )


    st.plotly_chart(fig, use_container_width = True)

    df_all



# Export data to Excel
if st.button("Export to Excel"):
    chemistries = df_all['Chemistry'].unique()

    # Initialize a new Workbook
    wb = Workbook()
    wb.remove(wb.active)  # remove the default sheet created by openpyxl

    for chem in chemistries:
        # Creating a list to store dataframes of each chemistry from different plates
        dfs_for_chem = []

        for plate_idx in range(num_plates):
            pandas_bio_df = bio_files[plate_idx]
            pandas_ise = ise_files[plate_idx]
            active_cells = all_active_cells[plate_idx]

            if pandas_bio_df is not None:
                bio_df_filtered = pandas_bio_df[(pandas_bio_df['Well Id'].isin(active_cells.keys())) & (pandas_bio_df['Chemistry'] == chem)]

                if not bio_df_filtered.empty:
                    dfs_for_chem.append(bio_df_filtered)

            if pandas_ise is not None:
                ise_df_filtered = pandas_ise[(pandas_ise['Well Id'].isin(active_cells.keys())) & (pandas_ise['Chemistry'] == chem)]

                if not ise_df_filtered.empty:
                    dfs_for_chem.append(ise_df_filtered)

        # Combine the dataframes in dfs_for_chem
        combined_df_for_chem = pd.concat(dfs_for_chem, ignore_index=True)

        # If we have any data for this chemistry, create a new sheet and write the data
        if not combined_df_for_chem.empty:
            ws = wb.create_sheet(title=chem)
            for r_idx, row in enumerate(combined_df_for_chem.iterrows(), 1):
                for c_idx, value in enumerate(row[1], 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
        wb.save(tmp_file.name)

        # Read the file into memory
        with open(tmp_file.name, "rb") as f:
            bytes_data = f.read()

    st.download_button(
        label="Download Excel file",
        data=bytes_data,
        file_name="YSI_Analyzer_Raw_Results.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )



